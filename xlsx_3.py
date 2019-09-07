import openpyxl, os
Path = 'C:\\Users\\蔡静静\Desktop'
os.chdir(Path)
Filename = 'python_example.xlsx'
wb = openpyxl.load_workbook(Filename)
sheet = wb['Sheet1']
#print(type(sheet))
print(sheet.max_row)
print(sheet.max_column)

from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1))

print(get_column_letter(100))

chsheet = wb['Sheet1']
print(get_column_letter(chsheet.max_column))
print(get_column_letter(chsheet.max_row))

print(column_index_from_string('AGD'))

print(tuple(sheet['A1:C2']))

for rowOfCellObject in sheet['A1:D7']:
    for CellObject in rowOfCellObject:
        print(CellObject.coordinate, CellObject.value)
    print('---End Of Row---')

sheet_2 = wb.active
list(sheet_2.columns)
print(sheet["B"])

for CellObj in sheet["C"]:
    print(CellObj.value)







