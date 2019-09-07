import openpyxl, os
filename = 'C:\\Users\\蔡静静\Desktop'
os.chdir(filename)
file = 'python_example.xlsx'
wb = openpyxl.load_workbook(file)

sheet = wb['Sheet1']
# sheet['A1']
ThatSheet = sheet['A1'].value
print(ThatSheet)

fruit = sheet['B1']
ThatFruit = fruit.value
print(ThatFruit)

cell_1 = sheet['C2']
cell_value = cell_1.value
print(cell_1)
print('The item in Row ' + str(cell_1.row) + ', Column ' + cell_1.column + ' is ' + cell_value)

print(sheet.cell(row=1, column=2))

print(sheet.cell(row=1, column=3).value)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=3).value)


