import openpyxl
from openpyxl import Workbook
wb = Workbook()

ws_1 = wb.create_sheet()
ws_2 = wb.create_sheet()

ws_titles = wb.sheetnames
print(ws_titles)

#wb.remove(wb['Sheet1'])
del wb['Sheet2']

new_wb = wb.sheetnames
print(new_wb)

sheet = wb['Sheet']
sheet['A1'] = 'Hello World'
print(sheet['A1'].value)

sheet_1 = wb['Sheet1']
sheet_1['A1'] = 200
sheet_1['A2'] = 300
sheet_1['A3'] = '=SUM(A1:A2)'

wb.save('create_xlsx.xlsx')

wbFormulas = openpyxl.load_workbook('create_xlsx.xlsx', data_only=True)
sheet_2 = wbFormulas['Sheet1']
print(sheet_2['A2'].value)

# wbDataOnly = openpyxl.load_workbook('create_xlsx.xlsx', data_only=True)
# Data = wbDataOnly['Sheet1']
# print(Data['A3'].value)

